from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import csv
from datetime import datetime


def generar_reporte_pdf(registros):
    """Genera un reporte en formato PDF"""
    response = HttpResponse(content_type='application/pdf')
    fecha = datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Accesos_{fecha}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1E88E5'),
        spaceAfter=30,
    )
    
    # Título
    elements.append(Paragraph('Reporte de Accesos al Edificio', title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Datos de la tabla
    data = [['Fecha y Hora', 'Nombre', 'Documento', 'Rol', 'Tipo', 'Estado', 'Punto de Acceso']]
    
    for registro in registros:
        nombre = registro.usuario.get_full_name() if registro.usuario else 'N/A'
        documento = registro.usuario.numero_identificacion if registro.usuario else 'N/A'
        rol = registro.credencial.rol.nombre if registro.credencial and registro.credencial.rol else 'N/A'
        
        data.append([
            registro.fecha_hora.strftime('%Y-%m-%d %H:%M:%S'),
            nombre,
            documento,
            rol,
            registro.tipo,
            registro.estado,
            registro.punto_acceso
        ])
    
    # Crear tabla
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E88E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    return response


def generar_reporte_excel(registros):
    """Genera un reporte en formato Excel"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    fecha = datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Accesos_{fecha}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Accesos"
    
    # Encabezados
    headers = ['Fecha y Hora', 'Nombre', 'Documento', 'Rol', 'Tipo', 'Estado', 'Punto de Acceso']
    ws.append(headers)
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color='1E88E5', end_color='1E88E5', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for registro in registros:
        nombre = registro.usuario.get_full_name() if registro.usuario else 'N/A'
        documento = registro.usuario.numero_identificacion if registro.usuario else 'N/A'
        rol = registro.credencial.rol.nombre if registro.credencial and registro.credencial.rol else 'N/A'
        
        ws.append([
            registro.fecha_hora.strftime('%Y-%m-%d %H:%M:%S'),
            nombre,
            documento,
            rol,
            registro.tipo,
            registro.estado,
            registro.punto_acceso
        ])
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(response)
    return response


def generar_reporte_csv(registros):
    """Genera un reporte en formato CSV"""
    response = HttpResponse(content_type='text/csv')
    fecha = datetime.now().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Accesos_{fecha}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Fecha y Hora', 'Nombre', 'Documento', 'Rol', 'Tipo', 'Estado', 'Punto de Acceso'])
    
    for registro in registros:
        nombre = registro.usuario.get_full_name() if registro.usuario else 'N/A'
        documento = registro.usuario.numero_identificacion if registro.usuario else 'N/A'
        rol = registro.credencial.rol.nombre if registro.credencial and registro.credencial.rol else 'N/A'
        
        writer.writerow([
            registro.fecha_hora.strftime('%Y-%m-%d %H:%M:%S'),
            nombre,
            documento,
            rol,
            registro.tipo,
            registro.estado,
            registro.punto_acceso
        ])
    
    return response




